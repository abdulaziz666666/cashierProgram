from tkinter import *
from tkinter.ttk import Treeview
import datetime
from tkinter.messagebox import showerror
import openpyxl
from pathlib import Path
from os import chdir
# import csv

WINDOW_WIDTH = 400
WINDOW_HEIGHT = 300
WINDOW_BG = '#7A8DAB'
todayDate = datetime.datetime.today()
currentYear = datetime.datetime.strftime(todayDate, '%Y')
currentMonth = datetime.datetime.strftime(todayDate, '%B')
currentDayNumber = datetime.datetime.strftime(todayDate, '%d')

btnColorStyles = {
    'add': {'bg': '#36D79A'},
    'delete': {'bg': '#E05851'},
    'process': {'bg': '#26B0DE'},
    'save': {'bg': '#CCC65C'}
}
btnGeometry = {'ipadx': 10, 'ipady': 3}


def setupTree(master, columns):
    columnsWithoutFirst = [c for c in columns if columns.index(c) != 0]

    tree = Treeview(master, columns=columnsWithoutFirst)
    tree.heading('#0', text=columns[0])
    tree.column('#0', anchor='center', width=120, stretch=False)

    for c in columnsWithoutFirst:
        tree.heading(c, text=c)
        tree.column(c, anchor='center', width=120, stretch=False)

    return tree

def resetEntries(*entries: Entry):
    for e in entries:
        e.delete(0, END)

def addProductToInvoice(numberEntry, nameEntry, priceEntry):
    global totalPrice

    try:
        totalPrice += float(priceEntry.get())

    except ValueError:
        showerror('خطأ', 'لم تقم بإدخال السعر')

    else:
        totalPriceLabel.config(text=str(totalPrice))
        invoiceTree.insert(
            '',
            'end',
            text=priceEntry.get(),
            values=(nameEntry.get(), numberEntry.get())
        )
        resetEntries(priceEntry, nameEntry, numberEntry)

def deleteProductFromInvoice():
    global totalPrice

    try:
        selectedProduct = invoiceTree.selection()[0]
        
    except IndexError:
        pass

    else:
        price = invoiceTree.item(selectedProduct, 'text')
        invoiceTree.delete(invoiceTree.selection()[0])

        totalPrice -= float(price)
        totalPriceLabel.config(text=str(totalPrice))


def saveChanges(selectedProduct, entries):
    global totalPrice
    priceEntry = entries.pop(0)
    invoiceTree.item(selectedProduct, text=priceEntry.get(), values=[e.get() for e in entries])

    try:
        newPrice = float(priceEntry.get())

    except ValueError:
        showerror('خطأ', 'لم تقم بإدخال السعر')

    else:
        oldPrice = float(selectionValues[0])
        totalPrice += newPrice - oldPrice
        totalPriceLabel.config(text=str(totalPrice))

        editWindow.destroy()

def editProductValues():
    global selectionValues, editWindow

    try:
        selectedProduct = invoiceTree.selection()[0]
        
    except IndexError:
        pass

    else:
        selectionValues = [invoiceTree.item(selectedProduct, 'text'),
                           *invoiceTree.item(selectedProduct, 'values')]

        editWindow = Toplevel(invoiceWindow)

        invoiceDataFrame = Frame(editWindow, bg=WINDOW_BG, width=300)
        invoiceDataFrame.grid(row=0, column=0, sticky='nsew')

        Label(invoiceDataFrame, text='رقم المنتج',
            bg=WINDOW_BG, fg='white', font=('', 12, 'bold')).grid(row=0, column=0, pady=(20, 10))
        Label(invoiceDataFrame, text='اسم المنتج', 
            bg=WINDOW_BG, fg='white', font=('', 12, 'bold')).grid(row=2, column=0, pady=10)
        Label(invoiceDataFrame, text='السعر', 
            bg=WINDOW_BG, fg='white', font=('', 12, 'bold')).grid(row=4, column=0, pady=10)
        
        numberEntry = Entry(invoiceDataFrame, width=30)
        nameEntry = Entry(invoiceDataFrame, width=30)
        priceEntry = Entry(invoiceDataFrame, width=30)
        saveChangesBtn = Button(invoiceDataFrame, btnColorStyles['process'], text='حفظ',
                         command=lambda: saveChanges(selectedProduct, [priceEntry, nameEntry, numberEntry]))

        numberEntry.grid(row=1, column=0, padx=40, sticky='we')
        nameEntry.grid(row=3, column=0, padx=40, sticky='we')
        priceEntry.grid(row=5, column=0, padx=40, sticky='we')
        saveChangesBtn.grid(btnGeometry, row=6, column=0, padx=39, pady=20, sticky='we')

        for e, v in zip((priceEntry, nameEntry, numberEntry), selectionValues):
            e.insert(0, v)

        editWindow.mainloop()



def saveRecordAtFile(values, currentTime):
    global totalPrice

    todayDirectory = Path(f'{currentYear}/{currentMonth}/{currentDayNumber}')
    todayDirectory.mkdir(parents=True, exist_ok=True)
    chdir(todayDirectory)

    workbook = openpyxl.Workbook()
    sheetName = datetime.datetime.strftime(currentTime, '%H;%M;%S')
    worksheet = workbook.create_sheet(sheetName, 0)
    
    for col, title in zip(range(1, 4), ('السعر', 'اسم المنتج', 'رقم المنتج')):
        worksheet.cell(row=1, column=col, value=title)

    lastRow = 0
    for row in range(2, len(values)+2):
        for valueIndex in range(1, 4):
            worksheet.cell(row=row, column=valueIndex, value=values[row-2][valueIndex-1])
        lastRow = row
    
    worksheet.cell(row=lastRow+1, column=1, value=str(totalPrice))
    
    workbook.save(f'{datetime.datetime.strftime(currentTime, '%H;%M;%S')}.xlsx')

def saveInvoiceAtRecord(*entries: Entry):
    global totalPrice
    currentTime = datetime.datetime.today()
    currentInvoiceSection = todayInvoicesTree.insert('', 'end', text=datetime.datetime.strftime(currentTime, '%H:%M:%S') + f' ({totalPrice} ريال)')

    invoiceItemsValues = [[invoiceTree.item(item, 'text'), *invoiceTree.item(item, 'values')] for item in invoiceTree.get_children()]
    for itemValues in invoiceItemsValues:
        itemValues = [f'({itemValues[0]} ريال)', itemValues[1]]
        todayInvoicesTree.insert(currentInvoiceSection, 'end', text=' '.join([v for v in itemValues]))
    
    saveRecordAtFile(invoiceItemsValues, currentTime)
    totalPrice = 0
    totalPriceLabel.config(text=str(totalPrice))
    resetEntries(*entries)
    invoiceTree.delete(*invoiceTree.get_children())

def createNewInvoice():
    global invoiceWindow, invoiceTree, totalPriceLabel, totalPrice
    totalPrice = 0

    invoiceWindow = Toplevel(window)
    invoiceWindow.title('إنشاء فاتورة')
    invoiceWindow.columnconfigure((0, 1, 2, 3, 4), weight=1)
    invoiceWindow.rowconfigure((0, 1, 2, 3, 4, 5, 6, 7, 8), weight=1)

    invoiceDataFrame = Frame(invoiceWindow, bg=WINDOW_BG, width=300)
    invoiceDataFrame.grid(row=0, column=3, rowspan=9, columnspan=2, sticky='nsew')

    Label(invoiceDataFrame, text='رقم المنتج',
          bg=WINDOW_BG, fg='white', font=('', 12, 'bold')).grid(row=0, column=3, columnspan=2, pady=(20, 10))
    Label(invoiceDataFrame, text='اسم المنتج', 
          bg=WINDOW_BG, fg='white', font=('', 12, 'bold')).grid(row=2, column=3, columnspan=2, pady=10)
    Label(invoiceDataFrame, text='السعر', 
          bg=WINDOW_BG, fg='white', font=('', 12, 'bold')).grid(row=4, column=3, columnspan=2, pady=10)
    
    numberEntry = Entry(invoiceDataFrame, width=30)
    nameEntry = Entry(invoiceDataFrame, width=30)
    priceEntry = Entry(invoiceDataFrame, width=30)
    addBtn = Button(invoiceDataFrame, btnColorStyles['add'], text='إضافة', command=lambda: addProductToInvoice(numberEntry, nameEntry, priceEntry))
    deleteBtn = Button(invoiceDataFrame, btnColorStyles['delete'], text='حذف', command=deleteProductFromInvoice)
    editBtn = Button(invoiceDataFrame, btnColorStyles['process'], text='تعديل', command=editProductValues)
    saveInvoiceBtn = Button(invoiceDataFrame, btnColorStyles['save'], text='حفظ الفاتورة', command=saveInvoiceAtRecord)

    numberEntry.grid(row=1, column=3, columnspan=2, padx=40, sticky='we')
    nameEntry.grid(row=3, column=3, columnspan=2, padx=40, sticky='we')
    priceEntry.grid(row=5, column=3, columnspan=2, padx=40, sticky='we')
    addBtn.grid(btnGeometry, row=6, column=3, padx=(40, 3), pady=(20, 0), sticky='we')
    deleteBtn.grid(btnGeometry, row=6, column=4, padx=(3, 40), pady=(20, 0), sticky='we')
    editBtn.grid(btnGeometry, row=7, column=3, columnspan=2, padx=40, pady=5, sticky='we')
    saveInvoiceBtn.grid(btnGeometry, row=8, column=3, columnspan=2, padx=40, pady=(0, 20), sticky='we')

    invoiceTree = setupTree(invoiceWindow, ('السعر', 'اسم المنتج', 'رقم المنتج'))
    invoiceTree.grid(row=0, column=0, rowspan=8, columnspan=3, sticky='nsew')

    totalPriceFrame = Frame(invoiceWindow, bg='lightgrey')
    totalPriceFrame.columnconfigure(0, weight=1)
    totalPriceFrame.rowconfigure(0, weight=1)
    totalPriceFrame.grid(row=8, column=0, columnspan=3, sticky='nsew')

    totalPriceLabel = Label(totalPriceFrame, text='0', font=('', 16, 'bold'))
    totalPriceLabel.grid(row=0, column=0, sticky='nsew')

    invoiceWindow.mainloop()

window = Tk()
window.title('برنامج المحاسبة')
window.columnconfigure(0, weight=1)
window.rowconfigure((0, 2), weight=1)
window.rowconfigure(1, weight=3)
window.config(bg=WINDOW_BG)

Label(window, text='سجل الفواتير اليوم ' + datetime.datetime.strftime(todayDate, '%Y/%m/%d'),
        font=('', 16, 'bold')).grid(row=0, column=0, sticky='nsew')

todayInvoicesTree = Treeview()
todayInvoicesTree.grid(row=1, column=0, sticky='nsew')

newInvoiceBtn = Button(window, btnColorStyles['add'], text='فاتورة جديدة', command=createNewInvoice)
newInvoiceBtn.grid(row=2, column=0, sticky='nsew')

window.mainloop()
